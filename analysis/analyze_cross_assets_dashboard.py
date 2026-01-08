# =============================================================================
# analyze_cross_assets_dashboard.py  (FASE 4B - Dashboard cross-asset)
#
# OBJETIVO
#   Construir un Excel "dashboard" a partir del Excel generado en Fase 4A
#   (analyze_cross_assets.py). El dashboard incluye:
#     - Resumen global: cantidad de assets/variables, cobertura, % missing
#     - Conteo de FLAGS (OK/REVISAR/CRITICO) y gráfico
#     - Rankings (Top N) y gráficos:
#         * missing_pct (peores)
#         * delta_abs_p95 (más "saltos")
#         * std / range_p95_p5 (variabilidad)
#         * bombas: starts_per_day_mean, on_pct (si existen)
#     - Pestañas por grupo (estanques/bombas/presion/macro) con Top N
#
# CRITERIOS / SUPUESTOS
#   - Este script NO recalcula indicadores: solo lee el Excel de Fase 4A.
#   - Usa la hoja "Indicators_All" como tabla maestra.
#   - Usa "Flags" para el estado operativo.
#   - Si algunas columnas no existen (ej bombas), el dashboard omite esos gráficos.
#
# EJEMPLOS DE USO (prompt/comando)
#   1) Generar dashboard desde el último fase4_cross:
#      python analysis/analyze_cross_assets_dashboard.py --input reports\\fase4_cross_20260108_120000.xlsx
#
#   2) Cambiar Top N y output:
#      python analysis/analyze_cross_assets_dashboard.py --input reports\\fase4_cross.xlsx --top 30 --output reports\\fase4_dashboard.xlsx
#
#   3) Solo algunos grupos:
#      python analysis/analyze_cross_assets_dashboard.py --input reports\\fase4_cross.xlsx --groups estanques bombas
# =============================================================================

import argparse
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


DEFAULT_GROUPS = ["estanques", "bombas", "presion", "macro_caudalimetro"]


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="FASE 4B: dashboard cross-asset (Excel con charts).")
    p.add_argument("--input", required=True, help="Excel generado por Fase 4A (fase4_cross_*.xlsx)")
    p.add_argument("--output", help="Excel de salida (default reports/fase4_dashboard_*.xlsx)")
    p.add_argument("--top", type=int, default=25, help="Top N para rankings y charts")
    p.add_argument("--groups", nargs="*", default=None, help=f"Grupos a incluir (default: {DEFAULT_GROUPS})")
    return p.parse_args()


def safe_sheet(name: str) -> str:
    name = str(name).strip() or "Sheet"
    bad = r"[]:*?/\\"
    for ch in bad:
        name = name.replace(ch, "_")
    return name[:31]


def autosize(ws, max_col: int, max_row: int, min_width: int = 10, max_width: int = 45) -> None:
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        best = min_width
        for row in range(1, max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            best = max(best, min(len(str(v)) + 2, max_width))
        ws.column_dimensions[letter].width = best


def header_style(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def write_table(ws, start_row: int, start_col: int, df: pd.DataFrame, title: str | None = None) -> tuple[int, int]:
    r = start_row
    c = start_col

    if title:
        ws.cell(row=r, column=c, value=title).font = Font(bold=True, size=12)
        r += 1

    # headers
    for j, colname in enumerate(df.columns, start=c):
        ws.cell(row=r, column=j, value=colname)
    header_style(ws, r, start_col + df.shape[1] - 1)
    r += 1

    # rows
    for _, rec in df.iterrows():
        for j, colname in enumerate(df.columns, start=c):
            val = rec[colname]
            # hacer floats más legibles
            if isinstance(val, float) and np.isfinite(val):
                ws.cell(row=r, column=j, value=float(val))
            else:
                ws.cell(row=r, column=j, value=val)
        r += 1

    end_row = r - 1
    end_col = c + df.shape[1] - 1
    return end_row, end_col


def add_barchart(ws, title: str, data_col: int, cat_col: int, header_row: int, first_data_row: int, last_row: int, anchor: str) -> None:
    """
    Bar chart simple (una serie) usando una columna de datos y una de categorías.
    """
    if last_row < first_data_row:
        return

    chart = BarChart()
    chart.title = title
    chart.height = 9
    chart.width = 18

    data = Reference(ws, min_col=data_col, min_row=header_row, max_row=last_row)
    cats = Reference(ws, min_col=cat_col, min_row=first_data_row, max_row=last_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, anchor)


def top_n(df: pd.DataFrame, metric: str, n: int, ascending: bool = False) -> pd.DataFrame:
    if metric not in df.columns:
        return pd.DataFrame()
    tmp = df.copy()
    tmp[metric] = pd.to_numeric(tmp[metric], errors="coerce")
    tmp = tmp[tmp[metric].notna()]
    if tmp.empty:
        return pd.DataFrame()
    cols = [c for c in ["group", "asset_id", "variable", metric] if c in tmp.columns]
    tmp = tmp.sort_values(metric, ascending=ascending).head(n)
    return tmp[cols]


def main() -> None:
    args = parse_args()
    in_path = Path(args.input)
    if not in_path.exists():
        raise SystemExit(f"[error] input no existe: {in_path}")

    groups = args.groups if args.groups else DEFAULT_GROUPS

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = Path(args.output) if args.output else Path("reports") / f"fase4_dashboard_{ts}.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # Leer hojas clave del excel fase 4A
    try:
        indicators = pd.read_excel(in_path, sheet_name="Indicators_All")
    except Exception as exc:
        raise SystemExit(f"[error] No pude leer sheet Indicators_All: {exc}") from exc

    flags = None
    try:
        flags = pd.read_excel(in_path, sheet_name="Flags")
    except Exception:
        flags = pd.DataFrame(columns=["group", "asset_id", "variable", "flag", "reason"])

    # Filtrar por grupos si aplica
    if "group" in indicators.columns:
        indicators = indicators[indicators["group"].astype(str).str.lower().isin([g.lower() for g in groups])].copy()
    if "group" in flags.columns and not flags.empty:
        flags = flags[flags["group"].astype(str).str.lower().isin([g.lower() for g in groups])].copy()

    # Resumen global
    n_rows = int(indicators.shape[0])
    n_assets = int(indicators["asset_id"].nunique()) if "asset_id" in indicators.columns else 0
    n_groups = int(indicators["group"].nunique()) if "group" in indicators.columns else 0

    # métricas globales
    missing_col = "missing_pct" if "missing_pct" in indicators.columns else None
    cont_col = "continuity_pct" if "continuity_pct" in indicators.columns else None

    global_missing_mean = float(pd.to_numeric(indicators[missing_col], errors="coerce").mean()) if missing_col else np.nan
    global_cont_mean = float(pd.to_numeric(indicators[cont_col], errors="coerce").mean()) if cont_col else np.nan

    # flags summary
    if flags is not None and not flags.empty and "flag" in flags.columns:
        flag_counts = flags["flag"].astype(str).value_counts().reindex(["OK", "REVISAR", "CRITICO"]).fillna(0).astype(int)
    else:
        flag_counts = pd.Series({"OK": 0, "REVISAR": 0, "CRITICO": 0})

    flags_df = pd.DataFrame({"flag": flag_counts.index, "count": flag_counts.values})

    # Rankings globales
    top = int(args.top)
    rankings = {
        "Top_missing_pct": top_n(indicators, "missing_pct", top, ascending=False),
        "Top_delta_abs_p95": top_n(indicators, "delta_abs_p95", top, ascending=False),
        "Top_std": top_n(indicators, "std", top, ascending=False),
        "Top_range_p95_p5": top_n(indicators, "range_p95_p5", top, ascending=False),
        "Top_pumps_starts_per_day": top_n(indicators[indicators.get("group", "").astype(str).str.lower() == "bombas"], "starts_per_day_mean", top, ascending=False),
        "Top_pumps_on_pct": top_n(indicators[indicators.get("group", "").astype(str).str.lower() == "bombas"], "on_pct", top, ascending=False),
    }

    # Crear workbook de salida (usamos openpyxl desde cero)
    # -> generamos un ExcelWriter y luego abrimos con openpyxl para agregar charts
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # Dashboard sheet (lo armamos manual luego con openpyxl, por ahora dejamos placeholder)
        pd.DataFrame().to_excel(writer, sheet_name="Dashboard", index=False)

        indicators.to_excel(writer, sheet_name="Indicators_All", index=False)
        flags.to_excel(writer, sheet_name="Flags", index=False)

        for name, df in rankings.items():
            if df is None or df.empty:
                continue
            df.to_excel(writer, sheet_name=safe_sheet(name), index=False)

        # por grupo
        if "group" in indicators.columns:
            for g, gdf in indicators.groupby(indicators["group"].astype(str), observed=True):
                gname = safe_sheet(f"Group_{g}")
                gdf.to_excel(writer, sheet_name=gname, index=False)

        # Notes
        notes = [
            ["source_fase4A", str(in_path)],
            ["included_groups", ", ".join(groups)],
            ["top_N", str(top)],
            ["generated", ts],
        ]
        pd.DataFrame(notes).to_excel(writer, sheet_name="Notes", index=False, header=False)

    # Abrir para escribir Dashboard + charts
    wb = load_workbook(out_path)
    ws = wb["Dashboard"]

    ws["A1"] = "FASE 4B - Dashboard Cross-Asset"
    ws["A1"].font = Font(bold=True, size=14)

    meta = pd.DataFrame(
        [
            ["Total filas (asset-variable)", n_rows],
            ["Total assets", n_assets],
            ["Total grupos", n_groups],
            ["Missing promedio (%)", global_missing_mean],
            ["Continuity promedio (%)", global_cont_mean],
        ],
        columns=["Métrica", "Valor"],
    )
    end_r, end_c = write_table(ws, 3, 1, meta, title="Resumen Global")

    # Flags table
    fr, fc = write_table(ws, end_r + 2, 1, flags_df, title="Conteo de FLAGS")
    # Chart flags
    # tabla flags: header en fila (end_r+2 + 1)
    flags_header_row = end_r + 3
    flags_first_data = flags_header_row + 1
    flags_last = fr
    add_barchart(
        ws,
        title="FLAGS (OK / REVISAR / CRITICO)",
        data_col=2,      # count
        cat_col=1,       # flag
        header_row=flags_header_row,
        first_data_row=flags_first_data,
        last_row=flags_last,
        anchor="E4",
    )

    # Sección Rankings en dashboard: ponemos 2 tablas lado a lado
    dash_row = fr + 2
    ws.cell(row=dash_row, column=1, value="Rankings Globales (Top N)").font = Font(bold=True, size=12)
    dash_row += 1

    blocks = [
        ("Top_missing_pct", "Peor Missing (%)", "missing_pct"),
        ("Top_delta_abs_p95", "Mayor DeltaAbs P95", "delta_abs_p95"),
        ("Top_std", "Mayor STD", "std"),
        ("Top_range_p95_p5", "Mayor Rango (P95-P5)", "range_p95_p5"),
    ]

    # Escribimos tablas una debajo de otra, con su chart a la derecha
    current_row = dash_row
    for key, title, metric in blocks:
        df = rankings.get(key)
        if df is None or df.empty:
            continue

        # Para chart: generamos label = asset_id|variable (más informativo)
        df2 = df.copy()
        if "asset_id" in df2.columns and "variable" in df2.columns:
            df2.insert(0, "label", df2["asset_id"].astype(str) + " | " + df2["variable"].astype(str))
        else:
            df2.insert(0, "label", df2.index.astype(str))

        # Solo label + metric para chart
        chart_df = df2[["label", metric]].copy()

        end_r2, end_c2 = write_table(ws, current_row, 1, chart_df, title=title)

        # chart
        header_row = current_row + 1
        first_data = header_row + 1
        last_row = end_r2
        add_barchart(
            ws,
            title=title,
            data_col=2,
            cat_col=1,
            header_row=header_row,
            first_data_row=first_data,
            last_row=last_row,
            anchor=f"E{current_row}",
        )

        current_row = end_r2 + 2

    # Bombas extras si hay
    pumps_blocks = [
        ("Top_pumps_starts_per_day", "Bombas: Arranques por día (mean)", "starts_per_day_mean"),
        ("Top_pumps_on_pct", "Bombas: % ON", "on_pct"),
    ]
    for key, title, metric in pumps_blocks:
        df = rankings.get(key)
        if df is None or df.empty:
            continue
        df2 = df.copy()
        df2.insert(0, "label", df2["asset_id"].astype(str))
        chart_df = df2[["label", metric]].copy()
        end_r2, _ = write_table(ws, current_row, 1, chart_df, title=title)

        header_row = current_row + 1
        first_data = header_row + 1
        add_barchart(
            ws,
            title=title,
            data_col=2,
            cat_col=1,
            header_row=header_row,
            first_data_row=first_data,
            last_row=end_r2,
            anchor=f"E{current_row}",
        )
        current_row = end_r2 + 2

    # formato/anchos
    max_row = max(ws.max_row, 50)
    max_col = max(ws.max_column, 10)
    autosize(ws, max_col=max_col, max_row=max_row)

    wb.save(out_path)
    print(f"[info] Fase 4B OK -> {out_path}")


if __name__ == "__main__":
    main()
