import argparse
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.chart import BarChart, Reference, ScatterChart, Series

TIMESTAMP_CANDIDATES = ["timestamp", "ts", "datetime", "fechahora"]
LEVEL_COLUMNS = {
    "nivelporcentual": "nivelPorcentual",
    "nivelestanque": "nivelEstanque",
}

MIN_AMPLITUDE_PCT = 10
MIN_DURATION_MIN = 30


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Analiza comportamiento de estanques y exporta reportes a Excel."
    )
    parser.add_argument("--input-dir", required=True, help="Carpeta base con CSVs")
    parser.add_argument("--asset", help="Filtra por asset (nombre o parte)")
    parser.add_argument(
        "--all",
        action="store_true",
        default=False,
        help="Procesa todos los assets detectados (default)",
    )
    parser.add_argument("--start-date", help="Fecha de inicio (YYYY-MM-DD)")
    parser.add_argument("--end-date", help="Fecha de fin (YYYY-MM-DD)")
    parser.add_argument("--freq-minutes", type=int, default=1, help="Frecuencia esperada")
    parser.add_argument("--output", help="Ruta de salida para el Excel (default reports/...)")
    return parser.parse_args()


def parse_timestamp(series: pd.Series) -> tuple[pd.Series, str]:
    numeric = pd.to_numeric(series, errors="coerce")
    if numeric.notna().any():
        median_value = float(numeric.dropna().median())
        if median_value >= 1e14:
            return pd.to_datetime(numeric, errors="coerce", unit="us"), "epoch_us"
        if median_value >= 1e11:
            return pd.to_datetime(numeric, errors="coerce", unit="ms"), "epoch_ms"
        if median_value >= 1e8:
            return pd.to_datetime(numeric, errors="coerce", unit="s"), "epoch_s"
    return pd.to_datetime(series, errors="coerce"), "string"


def parse_date(date_text: str | None, is_end: bool, freq_minutes: int) -> pd.Timestamp | None:
    if not date_text:
        return None
    parsed = pd.to_datetime(date_text, errors="coerce")
    if pd.isna(parsed):
        return None
    if is_end and len(date_text) == 10:
        parsed = parsed + pd.Timedelta(days=1) - pd.Timedelta(minutes=freq_minutes)
    return parsed


def find_timestamp_column(df: pd.DataFrame) -> str | None:
    lower_map = {col.lower(): col for col in df.columns}
    for candidate in TIMESTAMP_CANDIDATES:
        if candidate in lower_map:
            return lower_map[candidate]
    return None


def detect_level_columns(df: pd.DataFrame) -> dict[str, str]:
    lower_map = {col.lower(): col for col in df.columns}
    detected = {}
    for key, canonical in LEVEL_COLUMNS.items():
        if key in lower_map:
            detected[canonical] = lower_map[key]
    return detected


def determine_asset_series(df: pd.DataFrame, file_path: Path) -> pd.Series:
    lower_map = {col.lower(): col for col in df.columns}
    folder_name = file_path.parent.name
    prefix = file_path.stem.split("_")[0]
    fallback = folder_name or prefix
    asset_col = lower_map.get("asset_label")
    if asset_col:
        series = df[asset_col]
        if series.notna().any():
            series = series.astype(str).str.strip()
            series = series.where(series != "", fallback)
            return series
    return pd.Series([fallback] * len(df), index=df.index)


def load_csv(file_path: Path) -> tuple[pd.DataFrame | None, str | None, list[str]]:
    try:
        df = pd.read_csv(file_path)
    except Exception as exc:
        print(f"[warning] No se pudo leer {file_path}: {exc}")
        return None, None, []

    timestamp_col = find_timestamp_column(df)
    if not timestamp_col:
        print(f"[warning] Sin columna timestamp en {file_path}")
        return None, None, []

    df["timestamp"], strategy = parse_timestamp(df[timestamp_col])
    invalid_count = int(df["timestamp"].isna().sum())
    if invalid_count:
        print(f"[warning] {file_path} tiene {invalid_count} filas con timestamp inválido")
    df = df[df["timestamp"].notna()].copy()
    if df.empty:
        return None, strategy, []

    df["asset_id"] = determine_asset_series(df, file_path)
    level_columns = detect_level_columns(df)
    detected = list(level_columns.keys())
    for canonical, original in level_columns.items():
        df[canonical] = pd.to_numeric(df[original], errors="coerce")

    for canonical in LEVEL_COLUMNS.values():
        if canonical not in df.columns:
            df[canonical] = np.nan

    return df[["asset_id", "timestamp", *LEVEL_COLUMNS.values()]], strategy, detected


def compute_stats(series: pd.Series) -> dict[str, float]:
    total = int(series.shape[0])
    valid = int(series.notna().sum())
    missing_pct = (total - valid) / total * 100.0 if total else 0.0
    if valid == 0:
        return {
            "min": np.nan,
            "max": np.nan,
            "mean": np.nan,
            "median": np.nan,
            "std": np.nan,
            "n_total": total,
            "n_valid": valid,
            "missing_pct": missing_pct,
        }
    s = series.dropna()
    return {
        "min": float(s.min()),
        "max": float(s.max()),
        "mean": float(s.mean()),
        "median": float(s.median()),
        "std": float(s.std()),
        "n_total": total,
        "n_valid": valid,
        "missing_pct": missing_pct,
    }


def compute_percentiles(series: pd.Series, percentiles: list[int]) -> dict[int, float]:
    s = series.dropna()
    if s.empty:
        return {p: np.nan for p in percentiles}
    values = np.percentile(s.to_numpy(), percentiles)
    return {p: float(value) for p, value in zip(percentiles, values)}


def compute_histogram(series: pd.Series, bins: int = 20) -> pd.DataFrame:
    s = series.dropna()
    if s.empty:
        return pd.DataFrame(columns=["bin_left", "bin_right", "count"])
    counts, edges = np.histogram(s.to_numpy(), bins=bins)
    return pd.DataFrame(
        {
            "bin_left": edges[:-1],
            "bin_right": edges[1:],
            "count": counts,
        }
    )


def detect_recharge_events(asset_id: str, df: pd.DataFrame) -> list[dict]:
    data = df[["timestamp", "nivelPorcentual"]].dropna().sort_values("timestamp")
    if data.shape[0] < 3:
        return []
    smooth = data["nivelPorcentual"].rolling(window=5, center=True, min_periods=1).median()
    values = smooth.to_numpy()
    timestamps = data["timestamp"].to_numpy()
    minima = []
    maxima = []
    for idx in range(1, len(values) - 1):
        if values[idx - 1] > values[idx] <= values[idx + 1]:
            minima.append(idx)
        if values[idx - 1] < values[idx] >= values[idx + 1]:
            maxima.append(idx)

    events = []
    max_idx_iter = iter(sorted(maxima))
    try:
        current_max = next(max_idx_iter)
    except StopIteration:
        return []

    for min_idx in sorted(minima):
        while current_max <= min_idx:
            try:
                current_max = next(max_idx_iter)
            except StopIteration:
                return events
        start_time = pd.Timestamp(timestamps[min_idx])
        end_time = pd.Timestamp(timestamps[current_max])
        duration = (end_time - start_time).total_seconds() / 60.0
        amplitude = float(values[current_max] - values[min_idx])
        if amplitude >= MIN_AMPLITUDE_PCT and duration >= MIN_DURATION_MIN:
            events.append(
                {
                    "asset_id": asset_id,
                    "start_time": start_time,
                    "end_time": end_time,
                    "duration_minutes": duration,
                    "level_start": float(values[min_idx]),
                    "level_end": float(values[current_max]),
                    "amplitude": amplitude,
                }
            )
        try:
            current_max = next(max_idx_iter)
        except StopIteration:
            break
    return events


def compute_correlation(asset_id: str, df: pd.DataFrame) -> tuple[dict, pd.DataFrame]:
    pairs = df[["nivelPorcentual", "nivelEstanque"]].dropna()
    if pairs.shape[0] < 2:
        return {}, pd.DataFrame(columns=["asset_id", "nivelPorcentual", "nivelEstanque"])
    pearson_r = float(pairs["nivelPorcentual"].corr(pairs["nivelEstanque"]))
    slope, intercept = np.polyfit(
        pairs["nivelPorcentual"].to_numpy(), pairs["nivelEstanque"].to_numpy(), 1
    )
    predicted = slope * pairs["nivelPorcentual"] + intercept
    ss_res = float(((pairs["nivelEstanque"] - predicted) ** 2).sum())
    ss_tot = float(((pairs["nivelEstanque"] - pairs["nivelEstanque"].mean()) ** 2).sum())
    r2 = 1 - ss_res / ss_tot if ss_tot else np.nan
    summary = {
        "asset_id": asset_id,
        "pearson_r": pearson_r,
        "slope": float(slope),
        "intercept": float(intercept),
        "r2": float(r2),
    }
    scatter = pairs.copy()
    scatter.insert(0, "asset_id", asset_id)
    return summary, scatter


def add_summary_chart(ws, summary_df: pd.DataFrame) -> None:
    if summary_df.empty or "nivelPorcentual_missing_pct" not in summary_df.columns:
        return
    max_row = summary_df.shape[0] + 1
    asset_col = 1
    missing_col = summary_df.columns.get_loc("nivelPorcentual_missing_pct") + 1
    chart = BarChart()
    chart.title = "% Missing nivelPorcentual por asset"
    data = Reference(ws, min_col=missing_col, min_row=1, max_row=max_row)
    categories = Reference(ws, min_col=asset_col, min_row=2, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 8
    chart.width = 16
    ws.add_chart(chart, f"{chr(missing_col + 2)}2")


def add_recharge_chart(ws, start_row: int, bins_count: int) -> None:
    if bins_count <= 0:
        return
    chart = BarChart()
    chart.title = "Histograma amplitud recargas"
    data = Reference(ws, min_col=3, min_row=start_row, max_row=start_row + bins_count)
    categories = Reference(
        ws, min_col=1, min_row=start_row + 1, max_row=start_row + bins_count
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 8
    chart.width = 16
    ws.add_chart(chart, f"E{start_row}")


def add_correlation_chart(ws, start_row: int, data_rows: int) -> None:
    if data_rows <= 1:
        return
    chart = ScatterChart()
    chart.title = "nivelPorcentual vs nivelEstanque"
    chart.x_axis.title = "nivelPorcentual"
    chart.y_axis.title = "nivelEstanque"
    xvalues = Reference(ws, min_col=2, min_row=start_row + 1, max_row=start_row + data_rows)
    yvalues = Reference(ws, min_col=3, min_row=start_row + 1, max_row=start_row + data_rows)
    series = Series(yvalues, xvalues, title="Datos")
    chart.series.append(series)
    chart.height = 8
    chart.width = 16
    ws.add_chart(chart, f"E{start_row}")


def build_report(
    combined: pd.DataFrame,
    asset_filter: str | None,
    start_date: pd.Timestamp | None,
    end_date: pd.Timestamp | None,
    freq_minutes: int,
    output_path: Path,
    input_dir: Path,
    timestamp_strategies: list[str],
    detected_columns: dict[str, list[str]],
) -> None:
    summary_rows = []
    percentiles_rows = []
    hist_rows = []
    events_rows = []
    correlation_rows = []
    scatter_rows = []
    missing_notes = []

    filtered = combined.copy()
    if asset_filter:
        filtered = filtered[
            filtered["asset_id"].str.contains(asset_filter, case=False, na=False)
        ]

    percentiles = [1, 5, 10, 25, 50, 75, 90, 95, 99]

    for asset_id, asset_df in filtered.groupby("asset_id"):
        asset_df = asset_df.sort_values("timestamp")
        if start_date:
            asset_df = asset_df[asset_df["timestamp"] >= start_date]
        if end_date:
            asset_df = asset_df[asset_df["timestamp"] <= end_date]
        if asset_df.empty:
            print(f"[warning] Asset {asset_id} sin datos en el rango")
            continue

        date_min = asset_df["timestamp"].min()
        date_max = asset_df["timestamp"].max()

        summary = {"asset_id": asset_id, "date_min": date_min, "date_max": date_max}
        for canonical in LEVEL_COLUMNS.values():
            series = asset_df[canonical]
            stats = compute_stats(series)
            summary.update({f"{canonical}_{key}": value for key, value in stats.items()})
            perc = compute_percentiles(series, percentiles)
            for pct, value in perc.items():
                percentiles_rows.append(
                    {
                        "asset_id": asset_id,
                        "variable": canonical,
                        "percentile": f"P{pct}",
                        "value": value,
                    }
                )

            if canonical == "nivelPorcentual":
                hist_df = compute_histogram(series, bins=20)
                if not hist_df.empty:
                    hist_df.insert(0, "asset_id", asset_id)
                    hist_rows.append(hist_df)
                if series.dropna().empty:
                    missing_notes.append(f"{asset_id}: sin datos nivelPorcentual")

        summary_rows.append(summary)

        events_rows.extend(detect_recharge_events(asset_id, asset_df))

        if asset_df[["nivelPorcentual", "nivelEstanque"]].dropna().shape[0] >= 2:
            corr_summary, scatter = compute_correlation(asset_id, asset_df)
            if corr_summary:
                correlation_rows.append(corr_summary)
            if not scatter.empty:
                scatter_rows.append(scatter)

    summary_df = pd.DataFrame(summary_rows)
    percentiles_df = pd.DataFrame(percentiles_rows)
    hist_df = pd.concat(hist_rows, ignore_index=True) if hist_rows else pd.DataFrame(
        columns=["asset_id", "bin_left", "bin_right", "count"]
    )
    events_df = pd.DataFrame(events_rows)
    correlation_df = pd.DataFrame(correlation_rows)
    scatter_df = pd.concat(scatter_rows, ignore_index=True) if scatter_rows else pd.DataFrame(
        columns=["asset_id", "nivelPorcentual", "nivelEstanque"]
    )

    notes_rows = [
        ["input_dir", str(input_dir)],
        ["asset_filter", asset_filter or "--all"],
        ["start_date", str(start_date) if start_date else ""],
        ["end_date", str(end_date) if end_date else ""],
        ["freq_minutes", str(freq_minutes)],
        ["output", str(output_path)],
        ["MIN_AMPLITUDE_PCT", str(MIN_AMPLITUDE_PCT)],
        ["MIN_DURATION_MIN", str(MIN_DURATION_MIN)],
        ["timestamp_strategies", ", ".join(sorted(set(timestamp_strategies)))],
    ]
    for file_path, columns in detected_columns.items():
        notes_rows.append([f"columns:{file_path}", ", ".join(columns) or "(none)"])
    if missing_notes:
        notes_rows.append(["missing_nivelPorcentual", "; ".join(missing_notes)])

    notes_df = pd.DataFrame(notes_rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        percentiles_df.to_excel(writer, sheet_name="Percentiles", index=False)
        hist_df.to_excel(writer, sheet_name="Hist_nivelPorcentual", index=False)
        events_df.to_excel(writer, sheet_name="RechargeEvents", index=False)
        correlation_df.to_excel(writer, sheet_name="Correlation", index=False)
        start_row = len(correlation_df) + 3
        if not scatter_df.empty:
            scatter_df.to_excel(
                writer, sheet_name="Correlation", index=False, startrow=start_row
            )
        notes_df.to_excel(writer, sheet_name="Notes", index=False, header=False)

        summary_ws = writer.sheets["Summary"]
        add_summary_chart(summary_ws, summary_df)

        recharge_ws = writer.sheets["RechargeEvents"]
        if not events_df.empty:
            amplitudes = events_df["amplitude"].dropna()
            if not amplitudes.empty:
                bins_df = compute_histogram(amplitudes, bins=10)
                bins_start = events_df.shape[0] + 3
                recharge_ws.cell(row=bins_start, column=1, value="bin_left")
                recharge_ws.cell(row=bins_start, column=2, value="bin_right")
                recharge_ws.cell(row=bins_start, column=3, value="count")
                for idx, row in bins_df.iterrows():
                    recharge_ws.cell(row=bins_start + idx + 1, column=1, value=row["bin_left"])
                    recharge_ws.cell(row=bins_start + idx + 1, column=2, value=row["bin_right"])
                    recharge_ws.cell(row=bins_start + idx + 1, column=3, value=row["count"])
                add_recharge_chart(recharge_ws, bins_start, bins_df.shape[0])

        corr_ws = writer.sheets["Correlation"]
        if not scatter_df.empty:
            add_correlation_chart(corr_ws, start_row, scatter_df.shape[0])


def main() -> None:
    args = parse_args()
    input_dir = Path(args.input_dir)
    csv_paths = sorted(input_dir.rglob("*.csv"))
    if not csv_paths:
        print(f"[warning] No se encontraron CSVs en {input_dir}")

    data_frames = []
    timestamp_strategies = []
    detected_columns = {}

    for file_path in csv_paths:
        df, strategy, detected = load_csv(file_path)
        if df is None:
            continue
        data_frames.append(df)
        if strategy:
            timestamp_strategies.append(strategy)
        detected_columns[str(file_path)] = detected

    if data_frames:
        combined = pd.concat(data_frames, ignore_index=True)
    else:
        combined = pd.DataFrame(columns=["asset_id", "timestamp", *LEVEL_COLUMNS.values()])

    asset_filter = args.asset if args.asset else None
    start_date = parse_date(args.start_date, False, args.freq_minutes)
    end_date = parse_date(args.end_date, True, args.freq_minutes)

    if args.output:
        output_path = Path(args.output)
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = Path("reports") / f"tanks_{timestamp}.xlsx"

    build_report(
        combined,
        asset_filter,
        start_date,
        end_date,
        args.freq_minutes,
        output_path,
        input_dir,
        timestamp_strategies,
        detected_columns,
    )
    print(f"[info] Reporte generado en {output_path}")


if __name__ == "__main__":
    main()
